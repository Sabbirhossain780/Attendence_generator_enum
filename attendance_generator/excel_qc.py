import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


def build_qc_workbook(payloads: list, out_dir: str) -> str:
    COLOR_MAP = {
        "Survey":"C6EFCE","Training Day":"FFEB9C","Travel":"E2EFDA",
        "Weekend":"D9D9D9","Base work":"BDD7EE","":"FFFFFF",
    }
    all_dates = sorted(
        {r["date"] for p in payloads for r in p["rows"] if r["date"]},
        key=lambda s: pd.to_datetime(s)
    )

    wb    = openpyxl.Workbook()
    thin  = Side(style="thin", color="CCCCCC")
    bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)
    hfill = PatternFill("solid", fgColor="2E4057")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    ws = wb.active; ws.title="Matrix"
    c=ws.cell(1,1,"Enumerator"); c.font=hfont; c.fill=hfill
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr
    ws.column_dimensions["A"].width=28; ws.row_dimensions[1].height=36
    for ci,d in enumerate(all_dates,start=2):
        c=ws.cell(1,ci,d); c.font=hfont; c.fill=hfill; c.border=bdr
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width=13
    for ri,p in enumerate(payloads,start=2):
        d2p={r["date"]:r["purpose"] for r in p["rows"]}
        nc=ws.cell(ri,1,f"{p['enu_name']} ({p['enu_id']})")
        nc.font=Font(name="Arial",size=10); nc.border=bdr; nc.alignment=Alignment(vertical="center")
        for ci,d in enumerate(all_dates,start=2):
            purpose=d2p.get(d,""); c=ws.cell(ri,ci,purpose)
            c.font=Font(name="Arial",size=9); c.border=bdr
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.fill=PatternFill("solid",fgColor=COLOR_MAP.get(purpose,"FFFFFF"))
    ws.freeze_panes="B2"

    ws2=wb.create_sheet("Summary")
    hdrs=["ID","Name","Period Start","Period End","Data Days","Survey Days",
          "Training Days","Travel Days","Weekend Days","Base Work Days","Blank Days"]
    for ci,h in enumerate(hdrs,start=1):
        c=ws2.cell(1,ci,h); c.font=hfont; c.fill=hfill
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr
    for ri,p in enumerate(payloads,start=2):
        counts={}
        for r in p["rows"]: counts[r["purpose"]]=counts.get(r["purpose"],0)+1
        data_yes=sum(1 for r in p["rows"] if r["has_data"]=="Yes")
        vals=[p["enu_id"],p["enu_name"],p["period_start"],p["period_end"],
              data_yes,counts.get("Survey",0),counts.get("Training Day",0),
              counts.get("Travel",0),counts.get("Weekend",0),
              counts.get("Base work",0),counts.get("",0)]
        for ci,val in enumerate(vals,start=1):
            c=ws2.cell(ri,ci,val); c.font=Font(name="Arial",size=10); c.border=bdr
            c.alignment=Alignment(vertical="center")
            if ci==5: c.fill=PatternFill("solid",fgColor="C6EFCE")
    for ci in range(1,len(hdrs)+1):
        vals=[str(ws2.cell(r,ci).value or "") for r in range(1,len(payloads)+2)]
        ws2.column_dimensions[get_column_letter(ci)].width=min(max(len(v) for v in vals)+4,30)
    ws2.freeze_panes="A2"; ws2.row_dimensions[1].height=28

    ws3=wb.create_sheet("Legend")
    ws3.column_dimensions["A"].width=16; ws3.column_dimensions["B"].width=32
    ws3.cell(1,1,"Colour").font=Font(bold=True,name="Arial")
    ws3.cell(1,2,"Meaning").font=Font(bold=True,name="Arial")
    for i,(lbl,color) in enumerate([(k,v) for k,v in COLOR_MAP.items() if k],start=2):
        c=ws3.cell(i,1,lbl); c.fill=PatternFill("solid",fgColor=color); c.font=Font(name="Arial",size=10)
        ws3.cell(i,2,f"Day type: {lbl}").font=Font(name="Arial",size=10)

    qc_path=os.path.join(out_dir,"_QC_Summary.xlsx")
    wb.save(qc_path)
    return qc_path
